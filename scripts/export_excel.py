import sys
import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, Border, Alignment, PatternFill
from copy import copy

def copy_style(src_cell, dst_cell):
    """セルスタイルをコピーするヘルパー"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)

def export_excel(payload_path, template_path, output_path):
    # JSONペイロードを読み込み
    with open(payload_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # テンプレートを読み込み
    wb = openpyxl.load_workbook(template_path)
    if "申込書" not in wb.sheetnames:
        raise ValueError("申込書シートが見つかりません")
    
    ws = wb["申込書"]

    # --- ヘッダー情報の固定 (I3, P3) ---
    ROW_FACILITY = 3
    COL_FACILITY = 9 # I
    COL_DATE = 16 # P
    
    ws.cell(row=ROW_FACILITY, column=COL_FACILITY).value = f"施設名：{data.get('facilityName', '')}"
    
    year = data.get('year')
    month = data.get('month')
    day = data.get('day')
    if year and month and day:
        try:
            reiwa = int(year) - 2018
            # 形式: 施術日：令和 [年] 年 [月] 月 [日] 日 (スペース調整)
            ws.cell(row=ROW_FACILITY, column=COL_DATE).value = f"施術日：令和 {reiwa} 年 {month} 月 {day} 日"
        except:
            ws.cell(row=ROW_FACILITY, column=COL_DATE).value = f"施術日：令和    年   月   日"
    else:
        ws.cell(row=ROW_FACILITY, column=COL_DATE).value = f"施術日：令和    年   月   日"

    # --- 列構成の動的調整 ---
    # 元の固定列定義
    MENU_START_COL = 7        # G列
    TEMPLATE_MENU_COUNT = 6   # G-L列
    COL_TOTAL = 13            # M列
    COL_TIME_START = 14       # N列
    COL_SERVICE = 17          # Q列
    COL_REMARKS = 18          # R列
    
    ROW_MENU_HEADER = 9
    ROW_PRICE = 10
    ROW_DATA_START = 12
    
    customers = data.get('customers', [])
    
    # 追加項目の有無をチェック
    has_custom = any(c.get('isCustomOrder') for c in customers)
    has_guided = any(c.get('isGuided') for c in customers)
    has_additional = any(c.get('isAdditionalMenuAllowed') for c in customers)
    
    # 追加列の定義
    extra_cols = []
    if has_custom: extra_cols.append(('オーダーメイド', 'isCustomOrder', 0xFFE5CC)) # 薄オレンジ風
    if has_guided: extra_cols.append(('ご案内有無', 'isGuided', 0xCCE5FF)) # 薄青風
    if has_additional: extra_cols.append(('追加メニュー可否', 'isAdditionalMenuAllowed', 0xE5FFCC)) # 薄緑風

    # 列の挿入 (M列の前に挿入)
    num_inserted = 0
    for name, key, color in extra_cols:
        ws.insert_cols(COL_TOTAL + num_inserted)
        new_col = COL_TOTAL + num_inserted
        # ヘッダー設定
        cell = ws.cell(row=ROW_MENU_HEADER, column=new_col)
        cell.value = name
        # スタイルコピー (G列などのメニュー列からコピーして調整)
        copy_style(ws.cell(row=ROW_MENU_HEADER, column=MENU_START_COL), cell)
        if color:
            cell.fill = PatternFill(start_color=hex(color)[2:].zfill(6), end_color=hex(color)[2:].zfill(6), fill_type="solid")
        
        # 単価行は空に
        copy_style(ws.cell(row=ROW_PRICE, column=MENU_START_COL), ws.cell(row=ROW_PRICE, column=new_col))
        ws.cell(row=ROW_PRICE, column=new_col).value = ""
        
        num_inserted += 1

    # 列インデックスの更新
    COL_TOTAL += num_inserted
    COL_TIME_START += num_inserted
    COL_SERVICE += num_inserted
    COL_REMARKS += num_inserted

    # --- データの書き込み ---
    menu_items = data.get('menuItems', [])
    for i in range(min(len(menu_items), TEMPLATE_MENU_COUNT)):
        menu = menu_items[i]
        ws.cell(row=ROW_MENU_HEADER, column=MENU_START_COL + i).value = menu.get('name', '')
        ws.cell(row=ROW_PRICE, column=MENU_START_COL + i).value = menu.get('price', 0)

    for idx, customer in enumerate(customers):
        row_num = ROW_DATA_START + idx
        
        # 基本情報
        ws.cell(row=row_num, column=2).value = customer.get('no', idx + 1) # B
        ws.cell(row=row_num, column=3).value = customer.get('room', '')   # C
        ws.cell(row=row_num, column=4).value = customer.get('name', '')   # D
        ws.cell(row=row_num, column=6).value = customer.get('gender', '') # F

        # メニュー選択 (〇)
        selected_menus = customer.get('selectedMenus', [])
        for i in range(min(len(menu_items), TEMPLATE_MENU_COUNT)):
            menu_name = menu_items[i].get('name', '')
            ws.cell(row=row_num, column=MENU_START_COL + i).value = "〇" if menu_name in selected_menus else ""

        # 動的追加列のデータ
        current_extra_col = COL_TOTAL - num_inserted
        for name, key, color in extra_cols:
            val = customer.get(key, "")
            cell = ws.cell(row=row_num, column=current_extra_col)
            cell.value = val
            # スタイルを隣のセルからコピーして一貫性を保つ
            copy_style(ws.cell(row=row_num, column=MENU_START_COL), cell)
            current_extra_col += 1

        # 合計料金 (数式)
        formula_parts = []
        for i in range(min(len(menu_items), TEMPLATE_MENU_COUNT)):
            col_ltr = get_column_letter(MENU_START_COL + i)
            formula_parts.append(f'IF({col_ltr}{row_num}="〇",{col_ltr}{ROW_PRICE},0)')
        if formula_parts:
            ws.cell(row=row_num, column=COL_TOTAL).value = "=" + "+".join(formula_parts)

        # 希望時間
        times = customer.get('preferredTimes', [])
        for i in range(min(len(times), 3)):
            ws.cell(row=row_num, column=COL_TIME_START + i).value = times[i]

        # 施術実施
        if customer.get('hasService'):
            ws.cell(row=row_num, column=COL_SERVICE).value = "✓"

        # 備考
        ws.cell(row=row_num, column=COL_REMARKS).value = customer.get('remarks', '')

    # --- 合計人数行の追加 ---
    SUMMARY_ROW = ROW_DATA_START + len(customers)
    # ラベルセル (B-L列くらいまで結合されている場合が多いが、シンプルにBに書くか、適切な位置に)
    # ユーザーの「緑とか青」を再現するためにスタイルをヘッダーからコピー
    label_cell = ws.cell(row=SUMMARY_ROW, column=2) # B列
    label_cell.value = "合計人数"
    
    count_cell = ws.cell(row=SUMMARY_ROW, column=4) # D列あたりに人数
    count_cell.value = f"{len(customers)} 名"
    
    # 集計行のスタイル設定 (ヘッダーの色を模倣)
    header_style_cell = ws.cell(row=ROW_MENU_HEADER, column=2) # B列のヘッダー
    for c in range(2, COL_REMARKS + 1):
        target_cell = ws.cell(row=SUMMARY_ROW, column=c)
        copy_style(ws.cell(row=ROW_DATA_START, column=c), target_cell) # データ行のスタイル（罫線など）をコピー
        # 背景色を特定の色に（青っぽい色など）
        target_cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        target_cell.font = Font(bold=True)

    # --- 仕上げ ---
    for sheetname in wb.sheetnames:
        if sheetname != "申込書":
            del wb[sheetname]
    wb.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        sys.exit(1)
    try:
        export_excel(sys.argv[1], sys.argv[2], sys.argv[3])
        print("SUCCESS")
    except Exception as e:
        print(f"ERROR: {str(e)}", file=sys.stderr)
        sys.exit(1)
